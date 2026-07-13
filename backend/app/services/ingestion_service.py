"""Data ingestion pipeline: validate, clean, normalize, append to MASTER_DATASET."""

import io
import re
import uuid
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

import polars as pl

from app.config import Settings, get_settings
from app.domain.models import (
    FileUploadResult,
    UploadReport,
    UploadStatus,
    ValidationIssue,
    ValidationIssueType,
)
from app.domain.schema import (
    ALL_COLUMNS,
    BOOLEAN_COLUMNS,
    COLUMN_ALIASES,
    canonical_partner,
    date_header_priority,
    derive_campaign,
    derive_partner_from_source,
    FUNNEL_STAGE_RANK,
    FUNNEL_STAGES,
    FUZZY_COLUMN_RULES,
    CONTACT_STAGES_NORMALIZE,
    DATE_COLUMNS,
    DERIVED_COLUMNS,
    MASTER_PARQUET_FILE,
    NUMERIC_COLUMNS,
    OPTIONAL_COLUMNS,
    REQUIRED_COLUMNS,
    stage_rank,
)
from app.infrastructure.duckdb_repo import AnalyticsCache, DuckDBRepository
from app.logging_config import get_logger

logger = get_logger(__name__)


def normalize_header(header: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(header).strip().lower())
    if cleaned in COLUMN_ALIASES:
        return COLUMN_ALIASES[cleaned]
    # Google Sheets exports often suffix columns with "(Auto)" or "(auto)".
    without_auto = re.sub(r"\s*\(auto\)\s*$", "", cleaned).strip()
    if without_auto in COLUMN_ALIASES:
        return COLUMN_ALIASES[without_auto]
    # Keyword-based fallback for arbitrary naming variants
    # (e.g. "Prospect Creation Date", "Lead Creation Date", "Created On Date").
    for canon, needles in FUZZY_COLUMN_RULES:
        if all(n in without_auto for n in needles):
            return canon
    return without_auto.replace(" ", "_")


def apply_column_mapping(df: pl.DataFrame) -> pl.DataFrame:
    """Map workbook headers to canonical columns; coalesce when several map to one."""
    groups: Dict[str, List[str]] = {}
    for col in df.columns:
        canon = normalize_header(col)
        groups.setdefault(canon, []).append(col)

    exprs = []
    for canon, raws in groups.items():
        if len(raws) == 1:
            exprs.append(pl.col(raws[0]).alias(canon))
        elif canon == "date":
            ordered = sorted(raws, key=date_header_priority)
            exprs.append(
                pl.coalesce([pl.col(r).cast(pl.Utf8) for r in ordered]).alias(canon)
            )
        else:
            exprs.append(
                pl.coalesce([pl.col(r).cast(pl.Utf8) for r in raws]).alias(canon)
            )
    return df.select(exprs)


def normalize_partner(value: Optional[str]) -> Optional[str]:
    return canonical_partner(value)


def normalize_state(value: Optional[str]) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None
    return str(value).strip().title()


def normalize_contact_stage(value: Optional[str]) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None
    key = str(value).strip().lower()
    return CONTACT_STAGES_NORMALIZE.get(key, str(value).strip())


def normalize_phone(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    return digits if digits else None


def normalize_prospect_id(value: Any) -> Optional[str]:
    """Normalize Prospect ID from Excel (float/int/string) to a clean string."""
    if value is None:
        return None
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        if value == int(value):
            return str(int(value))
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return None
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


# Ordered date formats covering ISO, slash/dash numeric, and month-name styles
# (e.g. "8-May-2026", "8 May 2026", "May 8, 2026"). Google Sheets exports vary.
_DATE_FORMATS: Tuple[str, ...] = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%m-%d-%Y",
    "%d-%b-%Y",
    "%d-%B-%Y",
    "%d %b %Y",
    "%d %B %Y",
    "%b %d, %Y",
    "%B %d, %Y",
    "%b %d %Y",
    "%d-%b-%y",
)


# Windows Excel serial date epoch (1899-12-30).
_EXCEL_EPOCH = datetime(1899, 12, 30)
_EXCEL_SERIAL_MIN = 1_000
_EXCEL_SERIAL_MAX = 120_000


def _parse_excel_serial(value: float) -> Optional[datetime]:
    """Convert an Excel serial day number to datetime."""
    if not (_EXCEL_SERIAL_MIN <= value <= _EXCEL_SERIAL_MAX):
        return None
    whole = int(value)
    frac = value - whole
    dt = _EXCEL_EPOCH + timedelta(days=whole)
    if frac > 0:
        dt += timedelta(seconds=round(frac * 86400))
    return dt


def parse_date(value: Any) -> Optional[datetime]:
    """Parse a wide range of date representations into a datetime, or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:
            return None
        parsed = _parse_excel_serial(float(value))
        if parsed is not None:
            return parsed
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return None
    try:
        as_num = float(s.replace(",", ""))
        if re.fullmatch(r"-?\d+(?:\.\d+)?", s.replace(",", "")):
            parsed = _parse_excel_serial(as_num)
            if parsed is not None:
                return parsed
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # Last resort: let polars attempt inference on the raw string.
    try:
        parsed = pl.Series([s]).str.to_datetime(strict=False)
        val = parsed[0]
        return val if isinstance(val, datetime) else None
    except Exception:
        return None


def is_valid_date(value: Any) -> bool:
    """A blank date is allowed; any non-blank value must be parseable."""
    if value is None:
        return True
    if isinstance(value, (datetime, date)):
        return True
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return True
    return parse_date(s) is not None


def parse_bool(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ("1", "true", "yes", "y", "x", "✓", "connected", "admitted")


def parse_numeric(value: Any) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


_BOOL_TRUE_VALUES = ["1", "true", "yes", "y", "x", "✓", "connected", "admitted"]
_NAN_LIKE = ["", "nan", "none", "nat"]
ISSUE_SAMPLE_LIMIT = 50
NORMALIZE_CHUNK_SIZE = 100_000
MAX_DIAL_COUNT = 100

# Canonical dtype hints for MASTER_DATASET columns that are neither boolean nor
# in NUMERIC_COLUMNS (used when unifying schemas for the streaming parquet write).
_INT_MASTER_COLS = {"quarter", "year"}
_TS_MASTER_COLS = {"date", "ingested_at"}


def _build_failure_message(report: UploadReport) -> str:
    if report.rejection_summary:
        top = sorted(
            ((k, v) for k, v in report.rejection_summary.items() if v > 0),
            key=lambda x: x[1],
            reverse=True,
        )[:3]
        if top:
            parts = [f"{k.replace('_', ' ')}: {v:,}" for k, v in top]
            return f"No rows ingested — {', '.join(parts)}"
    if not report.issues:
        return "No rows were ingested"
    counts: Dict[str, int] = {}
    for issue in report.issues:
        key = issue.issue_type.value if hasattr(issue.issue_type, "value") else str(issue.issue_type)
        counts[key] = counts.get(key, 0) + 1
    top = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:3]
    parts = [f"{k.replace('_', ' ')}: {v:,}" for k, v in top]
    return f"No rows ingested — {', '.join(parts)}"


class IngestionEngine:
    """Incremental ingestion engine merging all workbooks into MASTER_DATASET."""

    def __init__(
        self,
        settings: Optional[Settings] = None,
        duck_repo: Optional[DuckDBRepository] = None,
        cache: Optional[AnalyticsCache] = None,
    ):
        self.settings = settings or get_settings()
        self.duck_repo = duck_repo or DuckDBRepository(self.settings)
        self.cache = cache or AnalyticsCache(self.settings.analytics_cache_ttl_seconds)
        self.parquet_path = self.settings.parquet_dir / MASTER_PARQUET_FILE

    def process_upload_batch(
        self,
        files: List[Tuple[str, bytes]],
        batch_id: Optional[str] = None,
        replace: bool = False,
        progress_cb: Optional[Callable[..., None]] = None,
    ) -> UploadReport:
        batch_id = batch_id or str(uuid.uuid4())
        started_at = datetime.utcnow()
        report = UploadReport(
            batch_id=batch_id,
            status=UploadStatus.PROCESSING,
            started_at=started_at,
            total_files=len(files),
        )

        def emit(percent: float, phase: str, rows_processed: int = 0, rows_total: int = 0) -> None:
            if progress_cb:
                try:
                    progress_cb(
                        percent=round(min(99.0, max(0.0, percent)), 1),
                        phase=phase,
                        rows_processed=rows_processed,
                        rows_total=rows_total,
                    )
                except Exception:
                    pass

        logger.info("upload_batch_started", batch_id=batch_id, file_count=len(files), replace=replace)
        emit(1, "Preparing")

        existing_ids = set() if replace else self.duck_repo.get_existing_prospect_ids()
        all_frames: List[pl.DataFrame] = []
        batch_seen_ids: Set[str] = set()
        duplicate_ids: Set[str] = set()

        expanded_files = self._expand_files(files)
        nfiles = max(len(expanded_files), 1)

        # Pass 1: read each file exactly once (records read errors immediately).
        emit(3, "Reading files")
        raw_items: List[Tuple[str, pl.DataFrame]] = []
        for i, (filename, content) in enumerate(expanded_files):
            ext = Path(filename).suffix.lower()
            if ext not in self.settings.allowed_extension_list:
                fr = FileUploadResult(filename=filename, success=False)
                fr.issues.append(
                    ValidationIssue(
                        issue_type=ValidationIssueType.INVALID_FILE_TYPE,
                        message=f"Unsupported file type: {ext}",
                    )
                )
                report.file_results.append(fr)
                report.issues.extend(fr.issues)
                continue
            try:
                df = self._read_raw_dataframe(filename, content)
            except Exception as exc:
                logger.error("file_read_error", filename=filename, error=str(exc))
                fr = FileUploadResult(filename=filename, success=False)
                fr.issues.append(
                    ValidationIssue(
                        issue_type=ValidationIssueType.CORRUPT_FILE,
                        message=f"Could not read file: {exc}",
                    )
                )
                report.file_results.append(fr)
                report.issues.extend(fr.issues)
                continue
            raw_items.append((filename, df))
            emit(3 + 7 * (i + 1) / nfiles, "Reading files")

        total_rows = sum(df.height for _, df in raw_items) or 1
        validated_rows = 0

        # Pass 2: validate + normalize each file (no second read of the source).
        for filename, df in raw_items:
            def row_progress(done_in_file: int) -> None:
                cur = validated_rows + done_in_file
                emit(10 + 55 * (cur / total_rows), "Validating rows", cur, total_rows)

            file_result, accepted_ids = self._validate_dataframe(
                df, filename, existing_ids, batch_seen_ids, duplicate_ids, progress=row_progress,
            )
            validated_rows += df.height
            report.file_results.append(file_result)
            report.total_rows_read += file_result.rows_read
            report.total_rows_accepted += file_result.rows_accepted
            report.total_rows_rejected += file_result.rows_rejected
            report.issues.extend(file_result.issues)

            if file_result.rows_accepted > 0:
                emit(68, "Normalizing rows", validated_rows, total_rows)

                def norm_progress(done_in_file: int, total_in_file: int) -> None:
                    emit(
                        68 + 20 * (done_in_file / max(total_in_file, 1)),
                        "Normalizing rows",
                        done_in_file,
                        total_in_file,
                    )

                frame = self._normalize_accepted(
                    df, filename, batch_id, accepted_ids, progress=norm_progress
                )
                if frame is not None and frame.height > 0:
                    all_frames.append(frame)
                    existing_ids.update(accepted_ids)

        report.duplicate_prospect_ids = sorted(duplicate_ids)
        report.duplicate_count = len(duplicate_ids)

        for fr in report.file_results:
            for issue_key, count in fr.rejection_summary.items():
                report.rejection_summary[issue_key] = (
                    report.rejection_summary.get(issue_key, 0) + count
                )

        if all_frames:
            emit(90, "Writing MASTER_DATASET")
            if replace:
                self.duck_repo.clear_master()
                self.cache.invalidate_all()
            self._write_master_batches(all_frames)
            emit(96, "Refreshing analytics")
            self.duck_repo.refresh_materialized_aggregates()
            self.cache.invalidate_all()
            report.status = UploadStatus.COMPLETED
            report.message = f"Successfully ingested {report.total_rows_accepted} rows into MASTER_DATASET"
        elif report.file_results and all(not fr.success for fr in report.file_results):
            report.status = UploadStatus.FAILED
            report.message = _build_failure_message(report)
        else:
            report.status = UploadStatus.PARTIAL if report.total_rows_accepted > 0 else UploadStatus.FAILED
            report.message = (
                f"Successfully ingested {report.total_rows_accepted} rows"
                if report.total_rows_accepted > 0
                else _build_failure_message(report)
            )

        report.completed_at = datetime.utcnow()
        report.master_dataset_total_rows = self.duck_repo.get_row_count()
        emit(99, "Finalizing")

        logger.info(
            "upload_batch_completed",
            batch_id=batch_id,
            status=report.status.value,
            rows_accepted=report.total_rows_accepted,
            master_total=report.master_dataset_total_rows,
        )
        return report

    def _expand_files(self, files: List[Tuple[str, bytes]]) -> List[Tuple[str, bytes]]:
        """Expand ZIP archives into individual workbook entries."""
        expanded: List[Tuple[str, bytes]] = []
        for filename, content in files:
            ext = Path(filename).suffix.lower()
            if ext == ".zip":
                try:
                    with zipfile.ZipFile(io.BytesIO(content)) as zf:
                        for name in zf.namelist():
                            if name.startswith("__MACOSX") or name.endswith("/"):
                                continue
                            inner_ext = Path(name).suffix.lower()
                            if inner_ext in (".xlsx", ".xls", ".csv"):
                                expanded.append((name, zf.read(name)))
                except zipfile.BadZipFile:
                    expanded.append((filename, content))
            else:
                expanded.append((filename, content))
        return expanded

    def _read_raw_dataframe(self, filename: str, content: bytes) -> pl.DataFrame:
        ext = Path(filename).suffix.lower()
        if ext == ".csv":
            return pl.read_csv(
                io.BytesIO(content),
                infer_schema_length=10000,
                ignore_errors=True,
                truncate_ragged_lines=True,
            )
        if ext in (".xlsx", ".xls"):
            return self._read_excel(content)
        raise ValueError(f"Unsupported extension: {ext}")

    def _read_excel(self, content: bytes) -> pl.DataFrame:
        """Read a workbook. Prefer calamine (fast, memory-efficient for large files);
        fall back to openpyxl with row padding to tolerate ragged rows."""
        try:
            sheets = pl.read_excel(
                io.BytesIO(content),
                sheet_id=0,  # 0 -> read every sheet, returns {name: DataFrame}
                engine="calamine",
                raise_if_empty=False,  # tolerate empty/blank sheets in the workbook
            )
            frames = [df for df in sheets.values() if df.width > 0 and df.height > 0]
            if frames:
                return pl.concat(frames, how="diagonal_relaxed")
            logger.warning("calamine_no_data_sheets_fallback_openpyxl")
        except Exception as exc:
            logger.warning("calamine_read_failed_fallback_openpyxl", error=str(exc))
        return self._read_excel_openpyxl(content)

    def _read_excel_openpyxl(self, content: bytes) -> pl.DataFrame:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets: List[pl.DataFrame] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            ncol = len(headers)
            # read_only mode drops trailing empty cells, producing ragged rows;
            # pad/truncate every row to the header width so polars can build columns.
            data = []
            for r in rows[1:]:
                r = list(r)
                if len(r) < ncol:
                    r.extend([None] * (ncol - len(r)))
                elif len(r) > ncol:
                    r = r[:ncol]
                data.append(r)
            schema = {h: pl.Utf8 for h in headers}
            sheets.append(pl.DataFrame(data, schema=schema, orient="row"))
        wb.close()
        if not sheets:
            raise ValueError("No data sheets found in workbook")
        return pl.concat(sheets, how="diagonal_relaxed")

    def _validate_dataframe(
        self,
        df: pl.DataFrame,
        filename: str,
        existing_ids: Set[str],
        batch_seen_ids: Set[str],
        duplicate_ids: Set[str],
        progress: Optional[Callable[[int], None]] = None,
    ) -> Tuple[FileUploadResult, Set[str]]:
        result = FileUploadResult(filename=filename)
        accepted_ids: Set[str] = set()
        result.rows_read = df.height

        df = apply_column_mapping(df)

        missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
        if missing:
            result.success = False
            for col in missing:
                result.issues.append(
                    ValidationIssue(
                        issue_type=ValidationIssueType.MISSING_COLUMN,
                        message=f"Required column missing: {col}",
                        column=col,
                    )
                )
            result.rows_rejected = result.rows_read
            result.rejection_summary = {
                ValidationIssueType.MISSING_COLUMN.value: result.rows_read
            }
            return result, accepted_ids

        # Vectorized validation: build boolean masks with native polars (parallel,
        # releases the GIL) instead of a per-row Python loop. Order of precedence
        # matches the original: blank id > duplicate > invalid date.
        seen = existing_ids | batch_seen_ids
        v = pl.DataFrame(
            {
                "_pid": self._pid_series(df),
                "_dparsed": self._parse_date_series(df),
                "_dblank": self._date_blank_series(df),
            }
        ).with_row_index("_row")

        v = v.with_columns(pl.col("_pid").is_null().alias("blank_pid"))
        v = v.with_columns(
            (~pl.col("blank_pid") & ~pl.col("_pid").is_first_distinct()).alias("dup_infile")
        )
        v = v.with_columns(
            (
                pl.col("_pid").is_in(list(seen)).fill_null(False)
                if seen
                else pl.lit(False)
            ).alias("dup_seen")
        )
        v = v.with_columns(
            (~pl.col("blank_pid") & (pl.col("dup_seen") | pl.col("dup_infile"))).alias("is_dup")
        )
        v = v.with_columns(
            (
                ~pl.col("blank_pid")
                & ~pl.col("is_dup")
                & ~pl.col("_dblank")
                & pl.col("_dparsed").is_null()
            ).alias("bad_date")
        )
        v = v.with_columns(
            (~pl.col("blank_pid") & ~pl.col("is_dup") & ~pl.col("bad_date")).alias("accepted")
        )

        agg = v.select(
            pl.col("blank_pid").sum().alias("n_blank"),
            pl.col("is_dup").sum().alias("n_dup"),
            pl.col("bad_date").sum().alias("n_baddate"),
            pl.col("accepted").sum().alias("n_ok"),
        ).row(0, named=True)

        n_blank = int(agg["n_blank"] or 0)
        n_dup = int(agg["n_dup"] or 0)
        n_baddate = int(agg["n_baddate"] or 0)
        n_ok = int(agg["n_ok"] or 0)

        accepted_ids = set(
            v.filter(pl.col("accepted")).get_column("_pid").to_list()
        )
        batch_seen_ids.update(accepted_ids)
        for pid_dup in v.filter(pl.col("is_dup")).get_column("_pid").unique().to_list():
            if pid_dup is not None:
                duplicate_ids.add(pid_dup)

        self._collect_issue_samples(result, v)
        result.rejection_summary = {
            k: n
            for k, n in (
                (ValidationIssueType.BLANK_PROSPECT_ID.value, n_blank),
                (ValidationIssueType.DUPLICATE_PROSPECT_ID.value, n_dup),
                (ValidationIssueType.INVALID_DATE.value, n_baddate),
            )
            if n
        }

        result.rows_accepted = n_ok
        result.rows_rejected = n_blank + n_dup + n_baddate
        if n_ok == 0 and result.rows_read > 0:
            result.success = False

        if progress:
            progress(result.rows_read)
        return result, accepted_ids

    # ---- Vectorized parsing helpers (shared by validate + normalize) ----

    @staticmethod
    def _pid_expr(col: str = "prospect_id") -> pl.Expr:
        """Normalize Prospect ID (float/int/str) to a clean string, blanks -> null."""
        s = pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
        s = s.str.replace(r"^(\d+)\.0+$", "${1}")
        low = s.str.to_lowercase()
        return pl.when(s.is_null() | low.is_in(_NAN_LIKE)).then(None).otherwise(s)

    def _pid_series(self, df: pl.DataFrame) -> pl.Series:
        return df.select(self._pid_expr().alias("_pid")).get_column("_pid")

    def _parse_date_series(self, df: pl.DataFrame) -> pl.Series:
        """Parse the date column dtype-aware. String dates are parsed by mapping
        only the distinct values through parse_date (dates repeat heavily)."""
        col = df.get_column("date")
        if col.dtype == pl.Datetime:
            return col
        if col.dtype == pl.Date:
            return col.cast(pl.Datetime)
        if col.dtype.is_numeric():
            uniques = col.unique().to_list()
            mapping = {u: parse_date(u) for u in uniques if u is not None}
            return df.select(
                pl.col("date")
                .replace_strict(mapping, default=None, return_dtype=pl.Datetime)
                .alias("_d")
            ).get_column("_d")
        s = col.cast(pl.Utf8, strict=False)
        uniques = s.unique().to_list()
        mapping = {u: parse_date(u) for u in uniques if u is not None}
        return df.select(
            pl.col("date")
            .cast(pl.Utf8, strict=False)
            .replace_strict(mapping, default=None, return_dtype=pl.Datetime)
            .alias("_d")
        ).get_column("_d")

    def _date_blank_series(self, df: pl.DataFrame) -> pl.Series:
        col = df.get_column("date")
        if col.dtype in (pl.Datetime, pl.Date):
            return col.is_null()
        s = col.cast(pl.Utf8, strict=False).str.strip_chars().str.to_lowercase()
        return (s.is_null() | s.is_in(_NAN_LIKE)).alias("_dblank")

    @staticmethod
    def _bool_expr(col: str) -> pl.Expr:
        return (
            pl.col(col)
            .cast(pl.Utf8, strict=False)
            .str.strip_chars()
            .str.to_lowercase()
            .is_in(_BOOL_TRUE_VALUES)
            .fill_null(False)
        )

    @staticmethod
    def _numeric_expr(col: str, dtype: pl.DataType) -> pl.Expr:
        if dtype.is_numeric():
            return pl.col(col).cast(pl.Float64, strict=False)
        return (
            pl.col(col)
            .cast(pl.Utf8, strict=False)
            .str.replace_all(",", "")
            .str.strip_chars()
            .cast(pl.Float64, strict=False)
        )

    def _map_unique(
        self, df: pl.DataFrame, col: str, func: Callable[[Any], Any], return_dtype=pl.Utf8
    ) -> pl.DataFrame:
        """Apply a Python normalizer to the DISTINCT values of a low-cardinality
        column, then map natively — avoids a per-row Python UDF over the whole frame."""
        if col not in df.columns:
            return df
        uniques = df.get_column(col).cast(pl.Utf8, strict=False).unique().to_list()
        mapping = {u: func(u) for u in uniques}
        return df.with_columns(
            pl.col(col)
            .cast(pl.Utf8, strict=False)
            .replace_strict(mapping, default=None, return_dtype=return_dtype)
            .alias(col)
        )

    def _derive_campaign_column(self, df: pl.DataFrame) -> pl.DataFrame:
        """Set campaign from data source batch (+ medium for College Wollege)."""
        batch_col = (
            pl.col("data_source_batch").cast(pl.Utf8, strict=False).str.strip_chars()
            if "data_source_batch" in df.columns
            else pl.lit(None).cast(pl.Utf8)
        )
        medium_col = (
            pl.col("medium").cast(pl.Utf8, strict=False).str.strip_chars()
            if "medium" in df.columns
            else pl.lit(None).cast(pl.Utf8)
        )
        partner_col = pl.col("partner").cast(pl.Utf8, strict=False)

        campaign_expr = (
            pl.when(partner_col == "College Wollege")
            .then(
                pl.when(batch_col.is_not_null() & (batch_col != "") & medium_col.is_not_null() & (medium_col != ""))
                .then(batch_col + pl.lit(" | ") + medium_col)
                .when(batch_col.is_not_null() & (batch_col != ""))
                .then(batch_col)
                .otherwise(medium_col)
            )
            .otherwise(
                pl.when(batch_col.is_not_null() & (batch_col != ""))
                .then(batch_col)
                .otherwise(None)
            )
        )

        if "campaign" in df.columns:
            df = df.with_columns(
                pl.when(campaign_expr.is_not_null())
                .then(campaign_expr)
                .otherwise(pl.col("campaign"))
                .alias("campaign")
            )
        else:
            df = df.with_columns(campaign_expr.alias("campaign"))
        return df

    def _collect_issue_samples(
        self, result: FileUploadResult, v: pl.DataFrame, sample: int = ISSUE_SAMPLE_LIMIT
    ) -> None:
        """Keep a small, representative set of example rejections (accurate totals
        live in rejection_summary) so the report stays small for huge files."""
        specs = [
            ("blank_pid", ValidationIssueType.BLANK_PROSPECT_ID, "Blank Prospect ID", "prospect_id"),
            ("is_dup", ValidationIssueType.DUPLICATE_PROSPECT_ID, "Duplicate Prospect ID", None),
            ("bad_date", ValidationIssueType.INVALID_DATE, "Invalid date", "date"),
        ]
        for mask_col, itype, msg, colname in specs:
            examples = v.filter(pl.col(mask_col)).select(["_row", "_pid"]).head(sample)
            for row in examples.iter_rows(named=True):
                is_dup = itype == ValidationIssueType.DUPLICATE_PROSPECT_ID
                result.issues.append(
                    ValidationIssue(
                        issue_type=itype,
                        message=f"{msg}: {row['_pid']}" if is_dup else msg,
                        row_number=int(row["_row"]) + 2,
                        column=colname,
                        prospect_id=row["_pid"],
                    )
                )

    def _normalize_accepted(
        self,
        df: pl.DataFrame,
        filename: str,
        batch_id: str,
        accepted_ids: Set[str],
        progress: Optional[Callable[[int, int], None]] = None,
    ) -> Optional[pl.DataFrame]:
        """Normalize accepted rows, batching large frames to bound peak memory
        and stream progress. Only accepted prospect IDs are kept."""
        try:
            df = apply_column_mapping(df)
            if accepted_ids:
                # Keep exactly the accepted rows, one per normalized Prospect ID
                # (a within-file duplicate shares the accepted id, so dedupe here).
                df = df.with_columns(self._pid_expr().alias("_npid"))
                df = df.filter(pl.col("_npid").is_in(list(accepted_ids)))
                df = df.unique(subset=["_npid"], keep="first", maintain_order=True).drop("_npid")
            if df.height == 0:
                return None

            if df.height <= NORMALIZE_CHUNK_SIZE:
                return self._normalize_dataframe(df, filename, batch_id)

            # Batch process very large frames in fixed-size chunks.
            parts: List[pl.DataFrame] = []
            done = 0
            for start in range(0, df.height, NORMALIZE_CHUNK_SIZE):
                chunk = df.slice(start, NORMALIZE_CHUNK_SIZE)
                parts.append(self._normalize_dataframe(chunk, filename, batch_id))
                done += chunk.height
                if progress:
                    progress(done, df.height)
            return pl.concat(parts, how="diagonal_relaxed")
        except Exception as exc:
            logger.error("normalize_error", filename=filename, error=str(exc))
            return None

    def _normalize_dataframe(
        self, df: pl.DataFrame, filename: str, batch_id: str
    ) -> pl.DataFrame:
        df = apply_column_mapping(df)

        for col in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
            if col not in df.columns:
                if col in BOOLEAN_COLUMNS:
                    df = df.with_columns(pl.lit(False).alias(col))
                elif col in NUMERIC_COLUMNS:
                    df = df.with_columns(pl.lit(None).cast(pl.Float64).alias(col))
                else:
                    df = df.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))

        # Prospect ID + text basics — native, vectorized.
        df = df.with_columns([
            self._pid_expr().alias("prospect_id"),
            pl.col("name").cast(pl.Utf8).str.strip_chars().alias("name"),
            pl.col("email").cast(pl.Utf8).str.strip_chars().str.to_lowercase().alias("email"),
        ])

        if "phone" in df.columns:
            df = df.with_columns(
                pl.col("phone").cast(pl.Utf8, strict=False).str.replace_all(r"\D", "").alias("phone")
            )
            df = df.with_columns(
                pl.when(pl.col("phone").str.len_chars() == 0)
                .then(None)
                .otherwise(pl.col("phone"))
                .alias("phone")
            )

        # Low-cardinality text columns: normalize only the distinct values in
        # Python, then map natively (avoids a per-row UDF over the whole frame).
        # Partner is derived from Contact Source, not a workbook Partner column.
        if "source" in df.columns:
            df = self._map_unique(
                df,
                "source",
                lambda v: str(v).strip() if v is not None and str(v).strip() else None,
            )
            src_uniques = df.get_column("source").unique().to_list()
            partner_map = {u: derive_partner_from_source(u) for u in src_uniques}
            df = df.with_columns(
                pl.col("source")
                .replace_strict(partner_map, default="Unknown", return_dtype=pl.Utf8)
                .alias("partner")
            )
        else:
            df = df.with_columns(pl.lit("Unknown").alias("partner"))

        df = self._derive_campaign_column(df)
        df = self._map_unique(df, "state", normalize_state)
        df = self._map_unique(df, "contact_stage", normalize_contact_stage)

        bool_exprs = [self._bool_expr(c).alias(c) for c in BOOLEAN_COLUMNS if c in df.columns]
        if bool_exprs:
            df = df.with_columns(bool_exprs)

        num_exprs = [
            self._numeric_expr(c, df.get_column(c).dtype).alias(c)
            for c in NUMERIC_COLUMNS
            if c in df.columns
        ]
        if num_exprs:
            df = df.with_columns(num_exprs)

        if "total_dialed_count" in df.columns:
            df = df.with_columns(
                pl.when(pl.col("total_dialed_count").fill_null(0) > MAX_DIAL_COUNT)
                .then(pl.lit(0))
                .otherwise(pl.col("total_dialed_count").fill_null(0).clip(0, MAX_DIAL_COUNT))
                .cast(pl.Int32)
                .alias("total_dialed_count")
            )

        df = df.with_columns(self._parse_date_series(df).alias("date"))

        df = self._add_derived_columns(df, filename, batch_id)

        select_cols = [c for c in ALL_COLUMNS if c in df.columns]
        return df.select(select_cols)

    def _add_derived_columns(
        self, df: pl.DataFrame, filename: str, batch_id: str
    ) -> pl.DataFrame:
        now = datetime.utcnow()

        df = df.with_columns([
            pl.when(pl.col("total_dialed_count").fill_null(0) == 0)
            .then(pl.lit("Never Dialed"))
            .when(pl.col("total_dialed_count") == 1)
            .then(pl.lit("1 Dial"))
            .when(pl.col("total_dialed_count") == 2)
            .then(pl.lit("2 Dial"))
            .otherwise(pl.lit("3+ Dial"))
            .alias("dial_bucket"),
        ])

        if "date" in df.columns:
            df = df.with_columns([
                pl.col("date").dt.strftime("%Y-W%V").alias("week"),
                pl.col("date").dt.strftime("%Y-%m").alias("month"),
                pl.col("date").dt.quarter().alias("quarter"),
                pl.col("date").dt.year().alias("year"),
                (pl.lit(now) - pl.col("date")).dt.total_days().alias("lead_age_days"),
            ])
        else:
            df = df.with_columns([
                pl.lit(None).cast(pl.Utf8).alias("week"),
                pl.lit(None).cast(pl.Utf8).alias("month"),
                pl.lit(None).cast(pl.Int32).alias("quarter"),
                pl.lit(None).cast(pl.Int32).alias("year"),
                pl.lit(None).cast(pl.Int32).alias("lead_age_days"),
            ])

        df = self._derive_funnel(df)

        df = self._derive_ai_from_contact_stage(df)

        # AI bot calling (excluding DNP / wrong number) counts as Connected.
        df = self._apply_ai_bot_connected(df)

        # After funnel flags are set, align contactability label with derived `connected`.
        df = df.with_columns([
            pl.when(pl.col("connected"))
            .then(pl.lit("Contactable"))
            .otherwise(pl.lit("Not Contactable"))
            .alias("contactability"),
        ])

        df = df.with_columns([
            pl.lit(now).alias("ingested_at"),
            pl.lit(filename).alias("source_file"),
            pl.lit(batch_id).alias("source_batch_id"),
        ])

        return df

    def _derive_ai_from_contact_stage(self, df: pl.DataFrame) -> pl.DataFrame:
        """Map Col D contact stages (AI Bot *) to ai_* boolean flags."""
        cs = (
            pl.col("contact_stage").cast(pl.Utf8, strict=False).str.to_lowercase()
            if "contact_stage" in df.columns
            else pl.lit("").cast(pl.Utf8)
        )
        has_ai = cs.str.contains("ai bot")

        flag_exprs = {
            "ai_contacted": has_ai,
            "ai_warm": has_ai & cs.str.contains("qualified - warm"),
            "ai_high_intent": has_ai & cs.str.contains("high intent"),
            "ai_payment_link": has_ai & cs.str.contains("payment link"),
            "ai_brochure": has_ai & cs.str.contains("brochure"),
            "ai_callback": has_ai & cs.str.contains("cb later"),
            "ai_interested": has_ai & (cs.str.contains("qualified - hot") | cs.str.contains("high intent")),
            "ai_qualified": has_ai
            & (
                cs.str.contains("qualified - warm")
                | cs.str.contains("qualified - hot")
                | cs.str.contains("high intent")
            ),
            "dnp": has_ai & cs.str.contains("reached - dnp"),
        }

        for col, expr in flag_exprs.items():
            if col in df.columns:
                df = df.with_columns((pl.col(col).fill_null(False) | expr).alias(col))
            else:
                df = df.with_columns(expr.alias(col))

        if "ai_status" in df.columns:
            df = df.with_columns(
                (
                    pl.col("ai_contacted")
                    | (pl.col("ai_status").is_not_null() & (pl.col("ai_status") != ""))
                ).alias("ai_contacted")
            )

        return df

    def _apply_ai_bot_connected(self, df: pl.DataFrame) -> pl.DataFrame:
        """Treat successful AI bot calling outcomes as Connected in the funnel.

        DNP and wrong-number outcomes stay at Lead (no meaningful connection).
        """
        if "ai_contacted" not in df.columns:
            return df

        cs = (
            pl.col("contact_stage").cast(pl.Utf8, strict=False).str.to_lowercase()
            if "contact_stage" in df.columns
            else pl.lit("")
        )
        ai_connected = (
            pl.col("ai_contacted").fill_null(False)
            & ~pl.col("dnp").fill_null(False)
            & ~cs.str.contains("wrong number")
        )

        df = df.with_columns(
            (pl.col("connected").fill_null(False) | ai_connected).alias("connected")
        )
        if "funnel_stage" in df.columns:
            df = df.with_columns(
                pl.when(
                    ai_connected
                    & (
                        pl.col("funnel_stage").is_null()
                        | (pl.col("funnel_stage") == "Lead")
                    )
                )
                .then(pl.lit("Connected"))
                .otherwise(pl.col("funnel_stage"))
                .alias("funnel_stage")
            )
        return df

    # Boolean flags that participate in the funnel, mapped to their rank.
    _FUNNEL_BOOL_RANK: Dict[str, int] = {
        "connected": FUNNEL_STAGE_RANK["Connected"],
        "mql": FUNNEL_STAGE_RANK["MQL"],
        "sql": FUNNEL_STAGE_RANK["SQL"],
        "application": FUNNEL_STAGE_RANK["Application"],
        "test_registration": FUNNEL_STAGE_RANK["Test Registration"],
        "interview": FUNNEL_STAGE_RANK["Interview"],
        "offer_letter": FUNNEL_STAGE_RANK["Offer Letter"],
        "block_amount_paid": FUNNEL_STAGE_RANK["Block Amount Paid"],
        "admission": FUNNEL_STAGE_RANK["Admission"],
    }

    def _derive_funnel(self, df: pl.DataFrame) -> pl.DataFrame:
        """Derive funnel_stage and cumulative stage flags from the furthest stage
        a prospect reached, considering the text lead/contact stages AND any
        pre-existing boolean flags. A prospect is always at least a "Lead"."""
        rank_exprs = [pl.lit(1, dtype=pl.Int32)]

        for col in ("lead_stage", "contact_stage"):
            if col in df.columns:
                distinct = df.select(pl.col(col)).unique().to_series().to_list()
                mapping = {v: stage_rank(v) for v in distinct if v is not None}
                if mapping:
                    rank_exprs.append(
                        pl.col(col)
                        .replace_strict(mapping, default=0, return_dtype=pl.Int32)
                        .fill_null(0)
                    )

        for col, rank in self._FUNNEL_BOOL_RANK.items():
            if col in df.columns:
                rank_exprs.append(
                    pl.when(pl.col(col).fill_null(False))
                    .then(pl.lit(rank, dtype=pl.Int32))
                    .otherwise(pl.lit(0, dtype=pl.Int32))
                )

        df = df.with_columns(pl.max_horizontal(*rank_exprs).alias("funnel_rank"))

        df = df.with_columns([
            (pl.col("funnel_rank") >= rank).alias(col)
            for col, rank in self._FUNNEL_BOOL_RANK.items()
        ])

        rank_to_stage = {FUNNEL_STAGE_RANK[s]: s for s in FUNNEL_STAGES}
        df = df.with_columns(
            pl.col("funnel_rank")
            .replace_strict(rank_to_stage, default="Lead", return_dtype=pl.Utf8)
            .alias("funnel_stage")
        )

        return df.drop("funnel_rank")

    def _append_to_master(self, new_data: pl.DataFrame) -> None:
        """Back-compat wrapper: write a single frame via the batched writer."""
        self._write_master_batches([new_data])

    @staticmethod
    def _canonical_dtype(col: str):
        if col in BOOLEAN_COLUMNS:
            return pl.Boolean
        if col in NUMERIC_COLUMNS:
            return pl.Float64
        if col in _INT_MASTER_COLS:
            return pl.Int32
        if col in _TS_MASTER_COLS:
            return pl.Datetime
        return pl.Utf8

    def _lazy_canonical(self, lf: pl.LazyFrame, present_cols: Set[str]) -> pl.LazyFrame:
        """Project a frame to the exact MASTER_DATASET schema (all columns, canonical
        dtypes, stable order) so heterogeneous sources concat + stream cleanly."""
        exprs = []
        for col in ALL_COLUMNS:
            if col in present_cols:
                exprs.append(pl.col(col).cast(self._canonical_dtype(col), strict=False).alias(col))
            else:
                exprs.append(pl.lit(None).cast(self._canonical_dtype(col)).alias(col))
        return lf.select(exprs)

    def _write_master_batches(self, frames: List[pl.DataFrame]) -> None:
        """Stream new (and any existing) rows into MASTER_DATASET in batches.

        Memory stays bounded: only a cheap per-partner count pass and the streaming
        parquet sink touch the whole dataset; nothing is fully materialized. The one
        global column, partner_share, is derived from the pre-computed totals.
        """
        frames = [f for f in frames if f is not None and f.height > 0]

        # Pass 1: global partner counts + grand total (existing + new).
        partner_counts: Dict[Any, int] = {}
        total = 0
        has_existing = self.parquet_path.exists()
        if has_existing:
            existing_counts = (
                pl.scan_parquet(self.parquet_path)
                .group_by("partner")
                .agg(pl.len().alias("n"))
                .collect()
            )
            for r in existing_counts.iter_rows(named=True):
                partner_counts[r["partner"]] = partner_counts.get(r["partner"], 0) + int(r["n"])
                total += int(r["n"])
        for f in frames:
            counts = f.group_by("partner").agg(pl.len().alias("n"))
            for r in counts.iter_rows(named=True):
                partner_counts[r["partner"]] = partner_counts.get(r["partner"], 0) + int(r["n"])
            total += f.height

        if total == 0:
            return
        denom = max(total, 1)
        share_map = {p: (n / denom * 100.0) for p, n in partner_counts.items()}

        # Pass 2: stream every source (existing parquet re-read in batches + each new
        # frame sliced into batches) through one incremental parquet writer. Only a
        # single batch is ever held/derived at a time, so peak memory stays flat.
        import pyarrow as pa
        import pyarrow.parquet as pq

        tmp_path = self.parquet_path.with_name(self.parquet_path.name + ".tmp")
        writer: Optional[pq.ParquetWriter] = None

        def write_batch(batch: pl.DataFrame) -> None:
            nonlocal writer
            table = self._finalize_batch(batch, share_map).to_arrow()
            if writer is None:
                writer = pq.ParquetWriter(str(tmp_path), table.schema, compression="zstd")
            else:
                table = table.cast(writer.schema)
            writer.write_table(table)

        try:
            if has_existing:
                # Open via a file handle so it is fully released before the swap
                # (Windows cannot replace a file that still has an open handle).
                with open(self.parquet_path, "rb") as handle:
                    parquet_file = pq.ParquetFile(handle)
                    for record_batch in parquet_file.iter_batches(batch_size=NORMALIZE_CHUNK_SIZE):
                        write_batch(pl.from_arrow(pa.Table.from_batches([record_batch])))
            for frame in frames:
                for start in range(0, frame.height, NORMALIZE_CHUNK_SIZE):
                    write_batch(frame.slice(start, NORMALIZE_CHUNK_SIZE))
        finally:
            if writer is not None:
                writer.close()

        tmp_path.replace(self.parquet_path)
        logger.info("master_dataset_updated", rows=total, path=str(self.parquet_path))

    def _finalize_batch(self, batch: pl.DataFrame, share_map: Dict[Any, float]) -> pl.DataFrame:
        """Project one batch to the canonical MASTER_DATASET schema and add the
        derived columns (global partner_share + row-local roi / conversion_pct)."""
        df = self._lazy_canonical(batch.lazy(), set(batch.columns)).collect()
        df = df.with_columns(
            pl.col("partner")
            .replace_strict(share_map, default=0.0, return_dtype=pl.Float64)
            .alias("partner_share")
        )
        df = df.with_columns(
            pl.when(pl.col("partner_cost").fill_null(0) > 0)
            .then((pl.col("revenue").fill_null(0) - pl.col("partner_cost").fill_null(0)) / pl.col("partner_cost"))
            .otherwise(pl.lit(0.0))
            .alias("roi")
        )
        df = df.with_columns(
            pl.when(pl.col("connected"))
            .then(
                pl.when(pl.col("admission")).then(pl.lit(100.0))
                .when(pl.col("offer_letter")).then(pl.lit(80.0))
                .when(pl.col("application")).then(pl.lit(60.0))
                .when(pl.col("mql")).then(pl.lit(40.0))
                .otherwise(pl.lit(20.0))
            )
            .otherwise(pl.lit(0.0))
            .alias("conversion_pct")
        )
        return df.select(ALL_COLUMNS)
