"""Block payment back-tracking sheet upload and storage."""

import io
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import polars as pl

from app.config import Settings, get_settings
from app.domain.schema import (
    BLOCK_PAYMENT_COLUMN_ALIASES,
    BLOCK_PAYMENT_COLUMNS,
    BLOCK_PAYMENT_META_FILE,
    BLOCK_PAYMENT_PARQUET_FILE,
)
from app.infrastructure.duckdb_repo import DuckDBRepository
from app.logging_config import get_logger
from app.services.ingestion_service import normalize_phone

logger = get_logger(__name__)


def normalize_block_payment_header(header: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(header).strip().lower())
    if cleaned in BLOCK_PAYMENT_COLUMN_ALIASES:
        return BLOCK_PAYMENT_COLUMN_ALIASES[cleaned]
    without_auto = re.sub(r"\s*\(auto\)\s*$", "", cleaned).strip()
    if without_auto in BLOCK_PAYMENT_COLUMN_ALIASES:
        return BLOCK_PAYMENT_COLUMN_ALIASES[without_auto]
    return without_auto.replace(" ", "_")


def normalize_match_email(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    email = str(value).strip().lower()
    return email if email and "@" in email else None


# Sheets: REGEXEXTRACT after application-fee"].*?utm_source "…" / utm_campaign "…"
_UTM_SOURCE_RE = re.compile(
    r'application-fee["\]]+.*?utm_source\s*"([^"]+)"',
    re.IGNORECASE | re.DOTALL,
)
_UTM_CAMPAIGN_RE = re.compile(
    r'application-fee["\]]+.*?utm_campaign\s*"([^"]+)"',
    re.IGNORECASE | re.DOTALL,
)


def _needs_utm_derive(value: Optional[str]) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return not text or text.lower() in {"not found", "none", "null", "nan"}


def extract_utm_source_at_payment(utm_activity: Optional[str]) -> Optional[str]:
    """Derive Source at Payment from Utm Activity (application-fee → utm_source)."""
    if utm_activity is None:
        return None
    text = str(utm_activity)
    if not text.strip():
        return None
    match = _UTM_SOURCE_RE.search(text)
    return match.group(1).strip() if match else None


def extract_utm_campaign_at_payment(utm_activity: Optional[str]) -> Optional[str]:
    """Derive Campaign at Payment from Utm Activity (application-fee → utm_campaign)."""
    if utm_activity is None:
        return None
    text = str(utm_activity)
    if not text.strip():
        return None
    match = _UTM_CAMPAIGN_RE.search(text)
    return match.group(1).strip() if match else None


def fill_payment_utm_from_activity(
    source_at_payment: Optional[str],
    campaign_at_payment: Optional[str],
    utm_activity: Optional[str],
) -> tuple[Optional[str], Optional[str]]:
    """Keep existing values; fill blanks / Not Found from utm_activity."""
    source = None if _needs_utm_derive(source_at_payment) else str(source_at_payment).strip()
    campaign = None if _needs_utm_derive(campaign_at_payment) else str(campaign_at_payment).strip()
    if source is None:
        source = extract_utm_source_at_payment(utm_activity)
    if campaign is None:
        campaign = extract_utm_campaign_at_payment(utm_activity)
    return source, campaign


def apply_block_payment_mapping(df: pl.DataFrame) -> pl.DataFrame:
    groups: Dict[str, List[str]] = {}
    for col in df.columns:
        canon = normalize_block_payment_header(col)
        groups.setdefault(canon, []).append(col)

    exprs = []
    for canon, raws in groups.items():
        if len(raws) == 1:
            exprs.append(pl.col(raws[0]).cast(pl.Utf8).alias(canon))
        else:
            exprs.append(
                pl.coalesce([pl.col(r).cast(pl.Utf8) for r in raws]).alias(canon)
            )
    return df.select(exprs)


class BlockPaymentService:
    """Upload and persist the block amount paid reconciliation sheet."""

    def __init__(
        self,
        duck_repo: Optional[DuckDBRepository] = None,
        settings: Optional[Settings] = None,
    ):
        self.settings = settings or get_settings()
        self.duck_repo = duck_repo or DuckDBRepository(self.settings)
        self.parquet_path = self.settings.parquet_dir / BLOCK_PAYMENT_PARQUET_FILE
        self.meta_path = self.settings.parquet_dir / BLOCK_PAYMENT_META_FILE

    def _read_file(self, filename: str, content: bytes) -> pl.DataFrame:
        ext = Path(filename).suffix.lower()
        if ext == ".csv":
            return pl.read_csv(
                io.BytesIO(content),
                infer_schema_length=10000,
                ignore_errors=True,
                truncate_ragged_lines=True,
            )
        if ext in (".xlsx", ".xls"):
            return self._read_excel(content)
        raise ValueError(f"Unsupported file type: {ext}. Use .xlsx, .xls, or .csv")

    def _read_excel(self, content: bytes) -> pl.DataFrame:
        try:
            sheets = pl.read_excel(
                io.BytesIO(content),
                sheet_id=0,
                engine="calamine",
                raise_if_empty=False,
            )
            frames = [df for df in sheets.values() if df.width > 0 and df.height > 0]
            if frames:
                return pl.concat(frames, how="diagonal_relaxed")
        except Exception as exc:
            logger.warning("block_payment_calamine_failed", error=str(exc))
        return self._read_excel_openpyxl(content)

    def _read_excel_openpyxl(self, content: bytes) -> pl.DataFrame:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets: List[pl.DataFrame] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            ncol = len(headers)
            data = []
            for r in rows[1:]:
                r = list(r)
                if len(r) < ncol:
                    r.extend([None] * (ncol - len(r)))
                elif len(r) > ncol:
                    r = r[:ncol]
                data.append(r)
            schema = {h: pl.Utf8 for h in headers}
            sheets.append(pl.DataFrame(data, schema=schema, orient="row"))
        wb.close()
        if not sheets:
            raise ValueError("Workbook has no data rows")
        return pl.concat(sheets, how="diagonal_relaxed")

    def _normalize_frame(self, df: pl.DataFrame, filename: str) -> pl.DataFrame:
        mapped = apply_block_payment_mapping(df)
        if mapped.height == 0:
            raise ValueError("Sheet has no data rows")

        for col in BLOCK_PAYMENT_COLUMNS:
            if col not in mapped.columns and col not in ("match_email", "match_phone", "uploaded_at", "source_filename"):
                mapped = mapped.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))

        uploaded_at = datetime.utcnow().isoformat()
        sources: List[Optional[str]] = []
        campaigns: List[Optional[str]] = []
        for row in mapped.select(
            ["source_at_payment", "campaign_at_payment", "utm_activity"]
        ).iter_rows(named=True):
            source, campaign = fill_payment_utm_from_activity(
                row.get("source_at_payment"),
                row.get("campaign_at_payment"),
                row.get("utm_activity"),
            )
            sources.append(source)
            campaigns.append(campaign)

        normalized = (
            mapped.with_columns(
                pl.Series("source_at_payment", sources, dtype=pl.Utf8),
                pl.Series("campaign_at_payment", campaigns, dtype=pl.Utf8),
                pl.col("email").map_elements(normalize_match_email, return_dtype=pl.Utf8).alias("match_email"),
                pl.col("phone").map_elements(normalize_phone, return_dtype=pl.Utf8).alias("match_phone"),
                pl.lit(uploaded_at).alias("uploaded_at"),
                pl.lit(filename).alias("source_filename"),
            )
            .select(BLOCK_PAYMENT_COLUMNS)
        )
        return normalized

    def upload_sheet(self, filename: str, content: bytes) -> Dict[str, Any]:
        if not content:
            raise ValueError("File is empty")

        raw = self._read_file(filename, content)
        frame = self._normalize_frame(raw, filename)
        self._reject_if_campus_fill_as_main(frame)
        row_count = frame.height

        tmp_path = self.parquet_path.with_suffix(".tmp.parquet")
        frame.write_parquet(tmp_path)
        tmp_path.replace(self.parquet_path)
        self.duck_repo.invalidate_metadata_cache()

        meta = {
            "uploaded_at": datetime.utcnow().isoformat(),
            "source_filename": filename,
            "row_count": row_count,
        }
        # Preserve prior campus-fill audit fields if present
        if self.meta_path.exists():
            try:
                prev = json.loads(self.meta_path.read_text(encoding="utf-8"))
                for key in (
                    "campus_fill_filename",
                    "campus_fill_uploaded_at",
                    "campus_fill_updated_count",
                ):
                    if key in prev:
                        meta[key] = prev[key]
            except Exception:
                pass
        self.meta_path.write_text(json.dumps(meta, indent=2), encoding="utf-8")

        logger.info("block_payment_sheet_uploaded", filename=filename, rows=row_count)
        return {
            "status": "completed",
            "row_count": row_count,
            "source_filename": filename,
            "uploaded_at": meta["uploaded_at"],
            "message": f"Uploaded {row_count} rows from {filename}",
        }

    def _nonempty_count(self, frame: pl.DataFrame, col: str) -> int:
        if col not in frame.columns:
            return 0
        return int(
            frame.filter(
                pl.col(col).is_not_null()
                & (pl.col(col).cast(pl.Utf8).str.strip_chars() != "")
            ).height
        )

    def _reject_if_campus_fill_as_main(self, frame: pl.DataFrame) -> None:
        """Block Amount Paid must keep gender/state/payment attrs — campus fill is a separate upload."""
        has_campus = self._nonempty_count(frame, "college_code") > 0
        has_gender = self._nonempty_count(frame, "gender") > 0
        has_state = self._nonempty_count(frame, "state") > 0
        has_payment_attrs = (
            self._nonempty_count(frame, "source_at_payment") > 0
            or self._nonempty_count(frame, "utm_activity") > 0
            or self._nonempty_count(frame, "campaign_at_payment") > 0
        )
        if has_campus and not has_gender and not has_state and not has_payment_attrs:
            raise ValueError(
                "This file looks like a campus fill sheet (CollegeCode/StudentEmail only). "
                "Upload the full Block Amount Paid Metabase export here first, then use the "
                "Campus fill dropzone below to patch blank campuses only — gender, state, and "
                "payment source stay on the main sheet."
            )

    def get_status(self) -> Dict[str, Any]:
        if not self.duck_repo.block_payment_exists():
            return {
                "has_data": False,
                "row_count": 0,
                "source_filename": None,
                "uploaded_at": None,
                "campus_fill_filename": None,
                "campus_fill_uploaded_at": None,
                "campus_fill_updated_count": 0,
                "blank_campus_count": 0,
            }

        meta: Dict[str, Any] = {}
        if self.meta_path.exists():
            try:
                meta = json.loads(self.meta_path.read_text(encoding="utf-8"))
            except Exception:
                meta = {}

        row_count = meta.get("row_count")
        if row_count is None:
            try:
                _, rows = self.duck_repo.execute_query(
                    f"SELECT COUNT(*) FROM read_parquet('{self.duck_repo._escape_path(self.parquet_path)}')"
                )
                row_count = int(rows[0][0]) if rows else 0
            except Exception:
                row_count = 0

        return {
            "has_data": True,
            "row_count": int(row_count),
            "source_filename": meta.get("source_filename"),
            "uploaded_at": meta.get("uploaded_at"),
            "campus_fill_filename": meta.get("campus_fill_filename"),
            "campus_fill_uploaded_at": meta.get("campus_fill_uploaded_at"),
            "campus_fill_updated_count": meta.get("campus_fill_updated_count"),
            "blank_campus_count": self._blank_campus_count(),
        }

    def _is_blank_campus(self, value: Any) -> bool:
        if value is None:
            return True
        text = str(value).strip()
        return not text or text.lower() in {"(blank)", "nan", "none", "null", "-"}

    def _blank_campus_count(self) -> int:
        if not self.parquet_path.exists():
            return 0
        try:
            frame = pl.read_parquet(self.parquet_path)
            if "college_code" not in frame.columns:
                return int(frame.height)
            return int(
                frame.filter(
                    pl.col("college_code").is_null()
                    | (pl.col("college_code").cast(pl.Utf8).str.strip_chars() == "")
                ).height
            )
        except Exception:
            return 0

    def list_blank_campus_rows(self) -> Dict[str, Any]:
        if not self.parquet_path.exists():
            return {"items": [], "total": 0}

        frame = pl.read_parquet(self.parquet_path)
        if frame.height == 0:
            return {"items": [], "total": 0}

        for col in ("email", "phone", "full_name", "gender", "college_code", "college_name", "sheet_id"):
            if col not in frame.columns:
                frame = frame.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))

        blank = frame.filter(
            pl.col("college_code").is_null()
            | (pl.col("college_code").cast(pl.Utf8).str.strip_chars() == "")
        )
        items = [
            {
                "sheet_id": r.get("sheet_id"),
                "email": r.get("email"),
                "phone": r.get("phone"),
                "full_name": r.get("full_name"),
                "gender": r.get("gender"),
                "college_code": r.get("college_code"),
                "college_name": r.get("college_name"),
            }
            for r in blank.select(
                ["sheet_id", "email", "phone", "full_name", "gender", "college_code", "college_name"]
            ).iter_rows(named=True)
        ]
        return {"items": items, "total": len(items)}

    def _normalize_campus_fill_frame(self, df: pl.DataFrame) -> pl.DataFrame:
        mapped = apply_block_payment_mapping(df)

        # Fallback for headers that slipped past aliases (e.g. mixed case CamelCase).
        rename_extra: Dict[str, str] = {}
        for col in list(mapped.columns):
            key = re.sub(r"[^a-z0-9]+", "", str(col).strip().lower())
            if key in {"campus", "collegecode"} and "college_code" not in mapped.columns:
                rename_extra[col] = "college_code"
            elif key in {"collegename", "campusname"} and "college_name" not in mapped.columns:
                rename_extra[col] = "college_name"
            elif key in {"studentemail", "email"} and "email" not in mapped.columns:
                rename_extra[col] = "email"
            elif key in {"studentphone", "phone"} and "phone" not in mapped.columns:
                rename_extra[col] = "phone"
            elif key in {"studentname", "fullname", "name"} and "full_name" not in mapped.columns:
                rename_extra[col] = "full_name"
        if rename_extra:
            mapped = mapped.rename(rename_extra)

        if "college_code" not in mapped.columns:
            raise ValueError(
                "Campus fill sheet needs a CollegeCode column "
                "(expected headers like StudentEmail, StudentPhone, CollegeCode, CollegeName)"
            )

        for col in ("email", "phone", "college_name"):
            if col not in mapped.columns:
                mapped = mapped.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))

        fill = mapped.with_columns(
            pl.col("email").map_elements(normalize_match_email, return_dtype=pl.Utf8).alias("match_email"),
            pl.col("phone").map_elements(normalize_phone, return_dtype=pl.Utf8).alias("match_phone"),
            pl.col("college_code").cast(pl.Utf8).str.strip_chars().alias("college_code"),
            pl.col("college_name").cast(pl.Utf8).str.strip_chars().alias("college_name"),
        ).filter(
            pl.col("college_code").is_not_null()
            & (pl.col("college_code") != "")
            & (
                (pl.col("match_email").is_not_null() & (pl.col("match_email") != ""))
                | (pl.col("match_phone").is_not_null() & (pl.col("match_phone") != ""))
            )
        )

        if fill.height == 0:
            raise ValueError(
                "Campus fill sheet has no usable rows "
                "(need StudentEmail/StudentPhone + CollegeCode)"
            )
        return fill.select(["match_email", "match_phone", "college_code", "college_name"])

    def apply_campus_fill_sheet(self, filename: str, content: bytes) -> Dict[str, Any]:
        if not content:
            raise ValueError("File is empty")
        if not self.parquet_path.exists():
            raise ValueError("Upload the Block Amount Paid sheet first")

        raw = self._read_file(filename, content)
        fill = self._normalize_campus_fill_frame(raw)
        block = pl.read_parquet(self.parquet_path)

        for col in BLOCK_PAYMENT_COLUMNS:
            if col not in block.columns:
                block = block.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))

        email_map: Dict[str, tuple[str, Optional[str]]] = {}
        phone_map: Dict[str, tuple[str, Optional[str]]] = {}
        for row in fill.iter_rows(named=True):
            code = str(row["college_code"]).strip()
            name = row.get("college_name")
            name_str = str(name).strip() if name is not None and str(name).strip() else None
            email = row.get("match_email")
            phone = row.get("match_phone")
            if email:
                email_map[str(email)] = (code, name_str)
            if phone:
                phone_map[str(phone)] = (code, name_str)

        codes: List[Optional[str]] = []
        names: List[Optional[str]] = []
        updated = 0
        matched_fill_keys: set[str] = set()

        for row in block.iter_rows(named=True):
            current_code = row.get("college_code")
            current_name = row.get("college_name")
            if not self._is_blank_campus(current_code):
                codes.append(None if current_code is None else str(current_code))
                names.append(None if current_name is None else str(current_name))
                continue

            email = row.get("match_email")
            phone = row.get("match_phone")
            hit: Optional[tuple[str, Optional[str]]] = None
            if email and str(email) in email_map:
                hit = email_map[str(email)]
                matched_fill_keys.add(f"e:{email}")
            elif phone and str(phone) in phone_map:
                hit = phone_map[str(phone)]
                matched_fill_keys.add(f"p:{phone}")

            if hit:
                codes.append(hit[0])
                names.append(hit[1] if hit[1] else (None if current_name is None else str(current_name)))
                updated += 1
            else:
                codes.append(None if current_code is None else str(current_code))
                names.append(None if current_name is None else str(current_name))

        patched = block.with_columns(
            pl.Series("college_code", codes, dtype=pl.Utf8),
            pl.Series("college_name", names, dtype=pl.Utf8),
        ).select(BLOCK_PAYMENT_COLUMNS)

        tmp_path = self.parquet_path.with_suffix(".tmp.parquet")
        patched.write_parquet(tmp_path)
        tmp_path.replace(self.parquet_path)
        self.duck_repo.invalidate_metadata_cache()

        fill_keys = set()
        for row in fill.iter_rows(named=True):
            if row.get("match_email"):
                fill_keys.add(f"e:{row['match_email']}")
            elif row.get("match_phone"):
                fill_keys.add(f"p:{row['match_phone']}")
        unmatched = max(0, len(fill_keys) - len(matched_fill_keys))

        still_blank = int(
            patched.filter(
                pl.col("college_code").is_null()
                | (pl.col("college_code").cast(pl.Utf8).str.strip_chars() == "")
            ).height
        )

        meta: Dict[str, Any] = {}
        if self.meta_path.exists():
            try:
                meta = json.loads(self.meta_path.read_text(encoding="utf-8"))
            except Exception:
                meta = {}
        filled_at = datetime.utcnow().isoformat()
        meta.update(
            {
                "row_count": patched.height,
                "campus_fill_filename": filename,
                "campus_fill_uploaded_at": filled_at,
                "campus_fill_updated_count": updated,
            }
        )
        self.meta_path.write_text(json.dumps(meta, indent=2), encoding="utf-8")

        logger.info(
            "block_payment_campus_fill_applied",
            filename=filename,
            updated=updated,
            unmatched=unmatched,
            still_blank=still_blank,
        )
        return {
            "status": "completed",
            "updated": updated,
            "unmatched": unmatched,
            "still_blank": still_blank,
            "source_filename": filename,
            "uploaded_at": filled_at,
            "message": (
                f"Filled campus on {updated} blank row(s) from {filename}"
                + (f"; {unmatched} fill row(s) unmatched" if unmatched else "")
                + (f"; {still_blank} still blank" if still_blank else "")
            ),
        }
