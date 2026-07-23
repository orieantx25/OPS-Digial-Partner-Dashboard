"""Persona last-24h activity report upload and storage."""

import io
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import polars as pl

from app.config import Settings, get_settings
from app.domain.schema import (
    PERSONA_ACTIVITY_COLUMN_ALIASES,
    PERSONA_ACTIVITY_COLUMNS,
    PERSONA_ACTIVITY_META_FILE,
    PERSONA_ACTIVITY_PARQUET_FILE,
)
from app.infrastructure.duckdb_repo import DuckDBRepository
from app.logging_config import get_logger
from app.services.block_payment_service import normalize_match_email

logger = get_logger(__name__)


def normalize_know_more_event_key(value: Any) -> str:
    """Normalize activity/persona text for Know More about B.Tech matching."""
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace(".", "")
        .replace(" ", "")
    )


def is_know_more_about_btech_event(notes: Any) -> bool:
    """True when activity EventName/notes is Know More about B.Tech."""
    return normalize_know_more_event_key(notes) == "knowmoreaboutbtech"


def normalize_persona_activity_header(header: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(header).strip().lower())
    if cleaned in PERSONA_ACTIVITY_COLUMN_ALIASES:
        return PERSONA_ACTIVITY_COLUMN_ALIASES[cleaned]
    without_auto = re.sub(r"\s*\(auto\)\s*$", "", cleaned).strip()
    if without_auto in PERSONA_ACTIVITY_COLUMN_ALIASES:
        return PERSONA_ACTIVITY_COLUMN_ALIASES[without_auto]
    return without_auto.replace(" ", "_")


def apply_persona_activity_mapping(df: pl.DataFrame) -> pl.DataFrame:
    groups: Dict[str, List[str]] = {}
    for col in df.columns:
        canon = normalize_persona_activity_header(col)
        groups.setdefault(canon, []).append(col)

    exprs = []
    for canon, raws in groups.items():
        if len(raws) == 1:
            exprs.append(pl.col(raws[0]).cast(pl.Utf8).alias(canon))
        else:
            exprs.append(
                pl.coalesce([pl.col(r).cast(pl.Utf8) for r in raws]).alias(canon)
            )
    return df.select(exprs)


class PersonaActivityService:
    """Upload and persist the persona activity (last 24h) report."""

    def __init__(
        self,
        duck_repo: Optional[DuckDBRepository] = None,
        settings: Optional[Settings] = None,
    ):
        self.settings = settings or get_settings()
        self.duck_repo = duck_repo or DuckDBRepository(self.settings)
        self.parquet_path = self.settings.parquet_dir / PERSONA_ACTIVITY_PARQUET_FILE
        self.meta_path = self.settings.parquet_dir / PERSONA_ACTIVITY_META_FILE

    def _read_file(self, filename: str, content: bytes) -> pl.DataFrame:
        ext = Path(filename).suffix.lower()
        if ext == ".csv":
            return pl.read_csv(
                io.BytesIO(content),
                infer_schema_length=10000,
                ignore_errors=True,
                truncate_ragged_lines=True,
            )
        if ext in (".xlsx", ".xls"):
            return self._read_excel(content)
        raise ValueError(f"Unsupported file type: {ext}. Use .xlsx, .xls, or .csv")

    def _read_excel(self, content: bytes) -> pl.DataFrame:
        try:
            sheets = pl.read_excel(
                io.BytesIO(content),
                sheet_id=0,
                engine="calamine",
                raise_if_empty=False,
            )
            frames = [df for df in sheets.values() if df.width > 0 and df.height > 0]
            if frames:
                return pl.concat(frames, how="diagonal_relaxed")
        except Exception as exc:
            logger.warning("persona_activity_calamine_failed", error=str(exc))
        return self._read_excel_openpyxl(content)

    def _read_excel_openpyxl(self, content: bytes) -> pl.DataFrame:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets: List[pl.DataFrame] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            ncol = len(headers)
            data = []
            for r in rows[1:]:
                r = list(r)
                if len(r) < ncol:
                    r.extend([None] * (ncol - len(r)))
                elif len(r) > ncol:
                    r = r[:ncol]
                data.append(r)
            schema = {h: pl.Utf8 for h in headers}
            sheets.append(pl.DataFrame(data, schema=schema, orient="row"))
        wb.close()
        if not sheets:
            raise ValueError("Workbook has no data rows")
        return pl.concat(sheets, how="diagonal_relaxed")

    def _normalize_frame(self, df: pl.DataFrame, filename: str) -> pl.DataFrame:
        mapped = apply_persona_activity_mapping(df)
        if mapped.height == 0:
            raise ValueError("Sheet has no data rows")

        required = {"prospect_id", "email"}
        present = set(mapped.columns)
        if not required.intersection(present):
            raise ValueError(
                "Persona report must include Prospect Id and/or Email Address columns"
            )

        for col in PERSONA_ACTIVITY_COLUMNS:
            if col not in mapped.columns and col not in (
                "match_email",
                "uploaded_at",
                "source_filename",
            ):
                mapped = mapped.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))

        uploaded_at = datetime.utcnow().isoformat()
        email_col = (
            pl.col("email")
            if "email" in mapped.columns
            else pl.lit(None).cast(pl.Utf8)
        )
        return (
            mapped.with_columns(
                email_col.map_elements(normalize_match_email, return_dtype=pl.Utf8).alias(
                    "match_email"
                ),
                pl.lit(uploaded_at).alias("uploaded_at"),
                pl.lit(filename).alias("source_filename"),
            )
            .select(PERSONA_ACTIVITY_COLUMNS)
        )

    def write_dataframe(
        self, frame: pl.DataFrame, source_label: str = "leadsquared_sync"
    ) -> Dict[str, Any]:
        """Persist a normalized persona activity frame (replaces prior upload)."""
        if frame is None:
            return {"row_count": 0, "message": "No persona activity rows to write"}

        uploaded_at = datetime.utcnow().isoformat()
        for col in PERSONA_ACTIVITY_COLUMNS:
            if col not in frame.columns:
                frame = frame.with_columns(pl.lit(None).cast(pl.Utf8).alias(col))
            else:
                frame = frame.with_columns(pl.col(col).cast(pl.Utf8, strict=False).alias(col))

        email_col = (
            pl.col("email") if "email" in frame.columns else pl.lit(None).cast(pl.Utf8)
        )
        frame = (
            frame.with_columns(
                email_col.map_elements(normalize_match_email, return_dtype=pl.Utf8).alias(
                    "match_email"
                ),
                pl.lit(uploaded_at).alias("uploaded_at"),
                pl.lit(source_label).alias("source_filename"),
            )
            .select(PERSONA_ACTIVITY_COLUMNS)
        )

        row_count = frame.height
        tmp_path = self.parquet_path.with_suffix(".tmp.parquet")
        frame.write_parquet(tmp_path)
        tmp_path.replace(self.parquet_path)

        meta = {
            "uploaded_at": uploaded_at,
            "source_filename": source_label,
            "row_count": row_count,
            "source": "leadsquared",
        }
        self.meta_path.write_text(json.dumps(meta, indent=2), encoding="utf-8")
        logger.info("persona_activity_synced", rows=row_count, source=source_label)
        return {
            "row_count": row_count,
            "source_filename": source_label,
            "uploaded_at": uploaded_at,
            "message": f"Synced {row_count} persona activity rows from LeadSquared",
        }

    def upload_sheet(self, filename: str, content: bytes) -> Dict[str, Any]:
        if not content:
            raise ValueError("File is empty")

        raw = self._read_file(filename, content)
        frame = self._normalize_frame(raw, filename)
        row_count = frame.height

        tmp_path = self.parquet_path.with_suffix(".tmp.parquet")
        frame.write_parquet(tmp_path)
        tmp_path.replace(self.parquet_path)

        meta = {
            "uploaded_at": datetime.utcnow().isoformat(),
            "source_filename": filename,
            "row_count": row_count,
        }
        self.meta_path.write_text(json.dumps(meta, indent=2), encoding="utf-8")

        logger.info("persona_activity_sheet_uploaded", filename=filename, rows=row_count)
        return {
            "status": "completed",
            "row_count": row_count,
            "source_filename": filename,
            "uploaded_at": meta["uploaded_at"],
            "message": (
                f"Uploaded {row_count} persona activity rows from {filename}. "
                "Last 24h charts now match these leads to the main dataset by Prospect Id / Email."
            ),
        }

    def get_status(self) -> Dict[str, Any]:
        if not self.duck_repo.persona_activity_exists():
            return {
                "has_data": False,
                "row_count": 0,
                "source_filename": None,
                "uploaded_at": None,
            }

        meta: Dict[str, Any] = {}
        if self.meta_path.exists():
            try:
                meta = json.loads(self.meta_path.read_text(encoding="utf-8"))
            except Exception:
                meta = {}

        row_count = meta.get("row_count")
        if row_count is None:
            try:
                _, rows = self.duck_repo.execute_query(
                    f"SELECT COUNT(*) FROM read_parquet('{self.duck_repo._escape_path(self.parquet_path)}')"
                )
                row_count = int(rows[0][0]) if rows else 0
            except Exception:
                row_count = 0

        return {
            "has_data": True,
            "row_count": int(row_count),
            "source_filename": meta.get("source_filename"),
            "uploaded_at": meta.get("uploaded_at"),
        }
